"""Excel (.xlsx/.xls) file parser using openpyxl."""

from __future__ import annotations

import logging
from pathlib import Path

from rag_kb.parsers.base import DocumentParser, ParsedDocument

logger = logging.getLogger(__name__)


class XlsxParser:
    """Parse Microsoft Excel spreadsheets, extracting all sheets as tables."""

    supported_extensions: list[str] = [".xlsx", ".xls"]

    def parse(self, file_path: Path) -> ParsedDocument:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel parsing: pip install openpyxl"
            ) from exc

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        sheets_text: list[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            header: list[str] = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(c).strip() if c is not None else "" for c in row]
                # Skip entirely empty rows
                if not any(cells):
                    continue

                if row_idx == 0:
                    header = cells
                    rows.append(" | ".join(cells))
                    rows.append(" | ".join("---" for _ in cells))
                else:
                    rows.append(" | ".join(cells))
                total_rows += 1

            if rows:
                sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

        wb.close()
        text = "\n\n".join(sheets_text)

        return ParsedDocument(
            text=text,
            source_path=str(file_path),
            metadata={
                "format": "xlsx",
                "sheet_count": str(len(wb.sheetnames)),
                "total_rows": str(total_rows),
            },
        )
